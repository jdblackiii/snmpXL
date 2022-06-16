from v3_tools import *
from v2_tools import *
from openpyxl import workbook
from openpyxl import load_workbook
from pysnmp import hlapi

# Basic order:
# > Read columns from workbook, basic iteration
# > SET functionality
# > GET functionality (req writing to workbook)
# > Package program to be launched from file, use cmd to specify
# > Allow specifying username, pass, etc from cmd
# > Implement a simple GUI

# Format excel document like so:
# > Column A: hostname or ip address
# > Column B: OID to modify or get
# > Column C: New value 
# Starts reading from line 2 so "headers" can be used

def setBulkv3(path, authProtocolIn, privProtocolIn, userIn, authPassIn, privPassIn):
    wb = load_workbook(filename = path)
    ws = wb.active
    currentIPCell = 'A2'
    currentOIDCell = 'B2'
    currrentValueCell = 'C2'
    rowNumber = 2
    #Iterates until we get a blank cell in A
    while ws[currentIPCell].value:
        setOnev3(ws[currrentValueCell].value, ws[currentOIDCell].value, ws[currentIPCell].value,
            authProtocolIn, privProtocolIn, userIn, authPassIn, privPassIn, 1)   
        #Set current cells to their next value
        rowNumber += 1 
        currentIPCell = 'A' + str(rowNumber)
        currentOIDCell = 'B' + str(rowNumber)
        currrentValueCell = 'C' + str(rowNumber)

# Gets the value of multiple OIDs and writes them to an excel document
# Format document as so: 
# IP , OID, Authentication password, privilege password, empty space for value
#
def getBulkv3(path, authProtocolIn, privProtocolIn, userIn = None, authPassIn = None, privPassIn = None):
    wb = load_workbook(filename = path)
    ws = wb.active
    currentIP = None
    currentOID = None
    currentAuthPass = authPassIn
    currentPrivPass = privPassIn
    currentUser = userIn
    currentValue = None
        
    rowNumber = 2
    
    #Loop while there are still values in the IP column
    while ws['A' + str(rowNumber)].value:
        currentIP = ws['A' + str(rowNumber)].value
        currentOID = ws['B' + str(rowNumber)].value

        if not currentAuthPass:
            currentAuthPass = ws['C' + str(rowNumber)].value

        if not currentPrivPass:
            currentPrivPass = ws['D' + str(rowNumber)].value

        if not currentUser:
            currentUser = ws['E' + str(rowNumber)].value
        
        currentValue = getOnev3(currentOID, currentIP,
            authProtocolIn, privProtocolIn, currentUser, currentAuthPass, currentPrivPass, 0)  
        
        if currentValue == 'authError':
            raise ValueError('IteratorCreationError', 'SNMP Value Constraint Error') 
        elif currentValue == 'otherError':
            raise ValueError('SNMPError', 'Error creating SNMP Object: Check credentials') 
        elif currentValue:
            ws['F' + str(rowNumber)] = currentValue
            
            print("Wrote value " + currentValue)
        else:
            print("Failed to write value at A{0}".format(rowNumber))
        
        rowNumber += 1 
            
    wb.save(path)
    print("Saved document at " + path)

def setBulkv2(path, communityString, toPrint):
    wb = load_workbook(filename = path)
    ws = wb.active
    currentIPCell = 'A2'
    currentOIDCell = 'B2'
    currrentValueCell = 'C2'
    rowNumber = 2
    #Iterates until we get a blank cell in A
    while ws[currentIPCell].value:
        setOnev2(ws[currrentValueCell].value, ws[currentIPCell].value, communityString, 
        ws[currentOIDCell].value, toPrint)   
        #Set current cells to their next value
        rowNumber += 1 
        currentIPCell = 'A' + str(rowNumber)
        currentOIDCell = 'B' + str(rowNumber)
        currrentValueCell = 'C' + str(rowNumber)


def getBulkv2(path, communityString, toPrint):
    wb = load_workbook(filename = path)
    ws = wb.active
    currentIPCell = 'A2'
    currentOIDCell = 'B2'
    currrentValueCell = 'C2'
        
    rowNumber = 2
    
    while ws[currentIPCell].value:           
            ws[currrentValueCell] = getOnev2(communityString, ws[currentIPCell].value, 
                ws[currentOIDCell].value, 0)  
            
            print("Wrote value " + ws[currrentValueCell].value + ' to ' + currrentValueCell)
            #Set current cells to their next value
            rowNumber += 1 
            currentIPCell = 'A' + str(rowNumber)
            currentOIDCell = 'B' + str(rowNumber)
            currrentValueCell = 'C' + str(rowNumber)

    wb.save(path)
    print("Saved document at " + path)