import sys
from pysnmp.hlapi import *
from v3_tools import *
from v2_tools import *
from openpyxl import Workbook, workbook
from openpyxl import load_workbook
from excel_integration import *

def callTest(ip, authProtocolIn, privProtocolIn, toPrint):
    return getOnev3('1.3.6.1.2.1.1.6.0', ip, authProtocolIn, privProtocolIn, 'jdLaptop', 'aeplongpass', 'aeplongpass', toPrint)

# discoverv3('1.3.6', '10.98.240.100', 
#         usmHMACMD5AuthProtocol, usm3DESEDEPrivProtocol,
#         'jdLaptop', 'aeplongpass', 'aeplongpass', 1)
# getWalkv3('10.98.240.100', usmHMACMD5AuthProtocol, usm3DESEDEPrivProtocol,
#         'jdLaptop', 'aeplongpass', 'aeplongpass')
# setOnev3('Lab', '1.3.6.1.2.1.1.6.0', '10.98.240.100', usmHMACMD5AuthProtocol, usm3DESEDEPrivProtocol, 
#     'jdLaptop', 'aeplongpass', 'aeplongpass', 1)
# callTest('10.98.240.100', usmHMACMD5AuthProtocol, usm3DESEDEPrivProtocol, 1)    #2690X
# callTest('10.7.1.3', usmHMACSHAAuthProtocol, usmAesCfb128Protocol, 1)           #ISA 3000
# callTest('10.98.228.189', usmHMACSHAAuthProtocol, usmAesCfb128Protocol, 1)      #LanTronix 8000
# getOnev2('SNMP-TEST', '10.98.240.100', '1.3.6.1.2.1.1.5.0')

# setBulk('C:\\Users\\S356430\\Documents\\VSCode\AEP SNMPpy\\test.xlsx', usmHMACMD5AuthProtocol, usm3DESEDEPrivProtocol, 'jdLaptop',
#     'aeplongpass', 'aeplongpass')

getOnev3('1.3.6.1.2.1.1.5.0', '10.7.1.3', usmHMAC128SHA224AuthProtocol, usmAesCfb128Protocol, 'V3_TEST',
    'aeplongpass', 'aeplongpass', 1)

# getBulkv3('C:\\Users\\S356430\\Documents\\VSCode\AEP SNMPpy\\test.xlsx', usmHMACSHAAuthProtocol, usmAesCfb128Protocol, userIn='S356430',
#     authPassIn='aeplongpass', privPassIn='aeplongpass')

# getOnev2('SNMP-TEST', '10.98.240.100', '1.3.6.1.2.1.1.5.0', 1)
# getBulkv2('C:\\Users\\S356430\\Documents\\VSCode\\AEP SNMPpy\\test.xlsx',
#    'SNMP-TEST', 1)

# wb = load_workbook(filename = 'C:\\Users\\S356430\\Documents\\VSCode\AEP SNMPpy\\test.xlsx')
# ws = wb.active
# print(ws['A2'].value)